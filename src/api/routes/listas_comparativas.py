"""Listas comparativas: listar, productos, upload Excel por proveedor."""
import io
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import JSONResponse
from pymongo.database import Database
from bson import ObjectId
from ..database import get_db

router = APIRouter()
COL = "LISTAS_COMPARATIVAS"


@router.get("/listas-comparativas/")
async def listar(db: Database = Depends(get_db)):
    col = db[COL]
    proveedores = db["PROVEEDORES"]
    listas = list(col.find({}))
    for L in listas:
        L["_id"] = str(L["_id"])
        try:
            prov = proveedores.find_one({"_id": ObjectId(L.get("proveedor_id", ""))})
            if prov:
                L["proveedor_empresa"] = prov.get("empresa")
        except Exception:
            pass
    return listas


@router.get("/listas-comparativas/productos")
async def productos(db: Database = Depends(get_db)):
    col = db[COL]
    proveedores = db["PROVEEDORES"]
    items = []
    for L in col.find({}):
        try:
            prov = proveedores.find_one({"_id": ObjectId(L.get("proveedor_id", ""))}) if L.get("proveedor_id") else None
        except Exception:
            prov = None
        desc = float(prov.get("condiciones_comerciales", 0) or 0) / 100.0 if prov else 0
        for p in L.get("productos", []):
            precio = float(p.get("precio", 0))
            precio_final = precio * (1 - desc) if desc else precio
            items.append({
                "proveedor_id": L.get("proveedor_id"),
                "proveedor_empresa": prov.get("empresa") if prov else "",
                "codigo": p.get("codigo"),
                "descripcion": p.get("descripcion"),
                "marca": p.get("marca"),
                "precio": precio,
                "precio_final": round(precio_final, 2),
                "existencia": p.get("existencia", 0),
            })
    return JSONResponse(content=items, status_code=200)


@router.post("/listas-comparativas/upload")
async def upload(file: UploadFile = File(...), proveedor_id: str = Form(...), db: Database = Depends(get_db)):
    try:
        prov = db["PROVEEDORES"].find_one({"_id": ObjectId(proveedor_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="proveedor_id inválido")
    if not prov:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    content = await file.read()
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl no instalado para leer Excel")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    productos = []
    headers = [str(c.value).lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = cell.value
        codigo = row_dict.get("codigo") or row_dict.get("Codigo")
        if not codigo:
            continue
        productos.append({
            "codigo": str(codigo),
            "descripcion": str(row_dict.get("descripcion") or row_dict.get("Descripcion") or ""),
            "marca": str(row_dict.get("marca") or row_dict.get("Marca") or ""),
            "precio": float(row_dict.get("precio") or row_dict.get("Precio") or 0),
            "existencia": int(row_dict.get("existencia") or row_dict.get("Existencia") or 0),
        })
    doc = {"proveedor_id": proveedor_id, "productos": productos}
    db[COL].insert_one(doc)
    return {"message": "Lista cargada", "productos_count": len(productos)}
